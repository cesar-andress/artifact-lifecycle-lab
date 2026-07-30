#!/usr/bin/env python3
"""Build blinded second-auditor package for all 121 RQ2 candidate events.

Does not invent human labels. Private answer key is written under
validation/rq2_second_audit/private/ and must stay out of public release
until independent annotation completes.
"""

from __future__ import annotations

import csv
import hashlib
import re
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover
    Workbook = None  # type: ignore

ROOT = Path(__file__).resolve().parents[2]
SRC = ROOT / "exports" / "truth_decay_pilot" / "rq2_failure_audit.csv"
OUT_DIR = ROOT / "validation" / "rq2_second_audit"
PRIVATE_DIR = OUT_DIR / "private"

# Fields that leak first-auditor conclusions or LLM judgments.
LEAK_FIELDS = frozenset(
    {
        "final_category",
        "category_letter",
        "is_genuine_decay",
        "adjudication_status",
        "heuristic_category",
        "heuristic_confidence",
        "heuristic_rules",
        "heuristic_rationale",
        "born_stale_heuristic_category",
        "born_stale_heuristic_confidence",
        "born_stale_heuristic_rules",
        "judge_a_model",
        "judge_a_category",
        "judge_a_rationale",
        "judge_b_model",
        "judge_b_category",
        "judge_b_rationale",
        "judge_agreement",
    }
)

EVIDENCE_FIELDS = (
    "event_id",
    "repo_id",
    "repo_url",
    "instruction_path",
    "reference_type",
    "reference",
    "time_origin",
    "time_end",
    "duration_days",
    "ever_repaired",
    "post_failure_followup_days",
    "failure_commit",
    "failure_transition",
    "verified_before_failure",
    "returned_after_missing",
    "basename_collision_verified",
    "n_observations",
    "repeated_repo_count",
    "repeated_file_count",
    "snippet_available",
    "snippet",
)

LABEL_FIELDS = (
    "event_id",
    "repo_id",
    "instruction_path",
    "reference_type",
    "reference",
    "failure_commit",
    "final_category",
    "category_letter",
    "is_genuine_decay",
    "adjudication_status",
    "heuristic_category",
    "heuristic_confidence",
    "heuristic_rules",
    "heuristic_rationale",
    "judge_a_category",
    "judge_b_category",
    "judge_agreement",
)

LEAK_PATTERNS = [
    re.compile(r"is_genuine_decay", re.I),
    re.compile(r"final_category", re.I),
    re.compile(r"genuine_decay", re.I),
    re.compile(r"0/121"),
    re.compile(r"heuristic_rationale", re.I),
    re.compile(r"judge_[ab]_category", re.I),
]


def event_id(row: dict[str, str]) -> str:
    key = "|".join(
        [
            row["repo_id"],
            row["instruction_path"],
            row["reference_type"],
            row["reference"],
            row["failure_commit"],
        ]
    )
    return hashlib.sha256(key.encode("utf-8")).hexdigest()[:16]


def load_rows() -> list[dict[str, str]]:
    with SRC.open(newline="", encoding="utf-8") as handle:
        rows = list(csv.DictReader(handle))
    if len(rows) != 121:
        raise SystemExit(f"expected 121 RQ2 audit rows, found {len(rows)}")
    return rows


def write_blinded_csv(rows: list[dict[str, str]], path: Path) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(EVIDENCE_FIELDS))
        writer.writeheader()
        for row in rows:
            out = {k: row.get(k, "") for k in EVIDENCE_FIELDS if k != "event_id"}
            out["event_id"] = event_id(row)
            # Reorder
            writer.writerow({k: out.get(k, row.get(k, "")) for k in EVIDENCE_FIELDS})


def write_private_key(rows: list[dict[str, str]], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(LABEL_FIELDS))
        writer.writeheader()
        for row in rows:
            out = {k: row.get(k, "") for k in LABEL_FIELDS if k != "event_id"}
            out["event_id"] = event_id(row)
            writer.writerow({k: out.get(k, "") for k in LABEL_FIELDS})


def write_xlsx(rows: list[dict[str, str]], path: Path) -> None:
    if Workbook is None:
        raise SystemExit("openpyxl required for audit form")
    wb = Workbook()
    ws = wb.active
    ws.title = "audit_form"
    headers = list(EVIDENCE_FIELDS) + [
        "auditor_category",
        "confidence",
        "counts_as_genuine_decay",
        "evidence_note",
        "category_ambiguity",
    ]
    header_fill = PatternFill("solid", fgColor="D9E2F3")
    for col, name in enumerate(headers, 1):
        cell = ws.cell(1, col, name)
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for r_i, row in enumerate(rows, 2):
        eid = event_id(row)
        for c_i, name in enumerate(EVIDENCE_FIELDS, 1):
            val = eid if name == "event_id" else row.get(name, "")
            cell = ws.cell(r_i, c_i, val)
            if name == "snippet":
                cell.alignment = Alignment(wrap_text=True)
        # annotation columns empty
        for c_i in range(len(EVIDENCE_FIELDS) + 1, len(headers) + 1):
            ws.cell(r_i, c_i, "")
    ws.freeze_panes = "A2"
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions["U"].width = 48  # snippet
    # codebook sheet with allowed values
    ws2 = wb.create_sheet("allowed_values")
    ws2["A1"] = "auditor_category"
    cats = [
        "genuine_decay",
        "rename_or_move",
        "verification_anchor_issue",
        "extractor_artifact",
        "normative_or_prescriptive",
        "external_or_environmental",
        "ambiguous",
        "uncertain_insufficient_evidence",
    ]
    for i, c in enumerate(cats, 2):
        ws2.cell(i, 1, c)
    ws2["B1"] = "confidence"
    for i, c in enumerate(["high", "medium", "low"], 2):
        ws2.cell(i, 2, c)
    ws2["C1"] = "counts_as_genuine_decay"
    for i, c in enumerate(["true", "false"], 2):
        ws2.cell(i, 3, c)
    wb.save(path)


def write_codebook(path: Path) -> None:
    path.write_text(
        """# RQ2 Second-Auditor Codebook

Use this taxonomy exactly. Do not invent new primary categories unless a case
cannot be represented; then use `uncertain_insufficient_evidence` and note why.

## Estimand reminder

You are classifying **first post-verification Missing** events: the reference was
Verified at least once, then later Missing under the frozen detector. Decide
whether the event is **genuine post-verification decay** for the RQ2 adjusted
estimand (physical disappearance / loss of the referenced artifact after prior
verified existence), or an alternative explanation.

You must **not** be told the original audit labels or the aggregate 0/121 result
while annotating.

## Categories

### A — `genuine_decay`

- **Definition:** After at least one verified observation, the referenced
  artifact genuinely disappears (or becomes unavailable) such that the
  instruction claim is now false for the same intended target.
- **Inclusion:** Clear before/after tree evidence of deletion or irreversible
  loss of the target; not explained by rename/move, extractor error, or
  normative/template text.
- **Exclusion:** Renames/moves with continuity; false detector triggers;
  anchors/templates; external URLs.
- **Minimum evidence:** Verified-before-failure plus failure commit evidence that
  the same target is gone without a rename/move account.
- **Counts as genuine decay for RQ2:** yes.

### B — `rename_or_move`

- **Definition:** Target identity continues under a new path/name (return after
  missing, basename collision with a verified peer, or clear move).
- **Inclusion:** `returned_after_missing` and/or rename/move evidence in history.
- **Exclusion:** True deletion with no continuity.
- **Minimum evidence:** Continuity signal (return or rename/move) tied to the
  same logical artifact.
- **Counts as genuine decay for RQ2:** no (under primary protocol).

### C — `verification_anchor_issue`

- **Definition:** Missing status driven by verification-anchor / path-resolution
  mismatch rather than physical disappearance of the intended artifact.
- **Inclusion:** Anchor or resolution mismatch explains the Missing transition.
- **Exclusion:** Unequivocal deletion of the intended target.
- **Minimum evidence:** Anchor/path evidence inconsistent with a simple delete.
- **Counts as genuine decay for RQ2:** no (under primary protocol).

### D — `extractor_artifact`

- **Definition:** The “reference” is an extraction/template/placeholder artifact
  (not a durable repository claim about a real path/URL).
- **Inclusion:** Placeholder patterns, extraction debris, non-claims.
- **Exclusion:** Real path/URL claims that later fail.
- **Minimum evidence:** Snippet or reference form showing non-claim / template.
- **Counts as genuine decay for RQ2:** no.

### E — `normative_or_prescriptive`

- **Definition:** Prescriptive/instructional language (“should”, “ensure”,
  examples) rather than an existential claim that a path currently exists.
- **Inclusion:** Normative framing in snippet.
- **Exclusion:** Factual existence claims.
- **Minimum evidence:** Snippet supporting normative reading.
- **Counts as genuine decay for RQ2:** no.

### F — `external_or_environmental`

- **Definition:** External resource / environment dependency outside repo tree
  verification scope.
- **Inclusion:** Clearly external target.
- **Exclusion:** In-repo paths.
- **Minimum evidence:** External URL/host or env marker.
- **Counts as genuine decay for RQ2:** no.

### G — `ambiguous`

- **Definition:** Competing explanations remain after reviewing available
  evidence; a category can be named but confidence is insufficient.
- **Inclusion:** Two plausible categories remain.
- **Exclusion:** Clear single category.
- **Minimum evidence:** Explicit competing accounts in the note.
- **Counts as genuine decay for RQ2:** no under primary protocol (sensitivity
  scenarios may reclassify).

### U — `uncertain_insufficient_evidence`

- **Definition:** Evidence package is insufficient to classify.
- **Inclusion:** Missing snippet/tree/history needed for a decision.
- **Exclusion:** Merely hard cases that still have enough evidence.
- **Counts as genuine decay for RQ2:** no (record separately).

## Required annotation fields

| Field | Values |
|-------|--------|
| `auditor_category` | A–G labels above, or `uncertain_insufficient_evidence` |
| `confidence` | `high` / `medium` / `low` |
| `counts_as_genuine_decay` | `true` / `false` — whether this event should count in the RQ2 adjusted numerator |
| `evidence_note` | concise note citing which evidence fields you used |
| `category_ambiguity` | optional: names of competing categories |

## Examples outside the 121 set

Use born-stale taxonomy documentation and non-RQ2 examples in
`exports/truth_decay_pilot/born_stale_examples.csv` and
`exports/truth_decay_pilot/gfc_confirmatory_examples.csv` for calibration.
Do **not** open `rq2_failure_audit.csv` or any private answer key while labeling.

## Blindedness

The blinded CSV/XLSX omit original `final_category`, `is_genuine_decay`,
heuristic conclusions, and LLM judge fields. Do not seek those files until
annotation is complete.
""",
        encoding="utf-8",
    )


def write_instructions(path: Path) -> None:
    path.write_text(
        """# Second-Auditor Instructions (RQ2, n=121)

## Role

You are an independent auditor. Classify each of the 121 blinded events using
`rq2_audit_codebook.md`. You should be independent of the original
classification. Prefer not to know the manuscript’s headline adjusted result
until after you finish.

## Materials

1. `rq2_audit_blinded.csv` — machine-readable evidence (no original labels).
2. `rq2_audit_form.xlsx` — same evidence plus empty annotation columns.
3. `rq2_audit_codebook.md` — category definitions and estimand.

## Procedure

1. Read the codebook once before labeling.
2. For each `event_id`, review evidence fields (reference, snippet, failure
   commit, verified-before-failure, returned-after-missing, basename collision,
   durations, repetition counts).
3. Fill `auditor_category`, `confidence`, `counts_as_genuine_decay`,
   `evidence_note`, and `category_ambiguity` if needed.
4. Do not leave rows blank; use `uncertain_insufficient_evidence` when stuck.
5. Save your completed sheet as:

   `rq2_second_auditor_labels.csv`

   with columns:

   `event_id,auditor_category,confidence,counts_as_genuine_decay,evidence_note,category_ambiguity,auditor_id,annotation_date`

6. Do not open `private/rq2_original_labels_private.csv` or the frozen
   `exports/truth_decay_pilot/rq2_failure_audit.csv` until labeling is done.

## After labeling

Return the completed CSV to the analysis maintainer. Agreement, kappa, and
adjudication scripts will then be run. Disagreements that change the binary
estimand will be adjudicated with an explicit rationale file; original labels
are never silently overwritten.
""",
        encoding="utf-8",
    )


def assert_no_leakage(blinded_path: Path) -> None:
    text = blinded_path.read_text(encoding="utf-8")
    # header check
    header = text.splitlines()[0]
    for field in LEAK_FIELDS:
        if field in header.split(","):
            raise SystemExit(f"leakage: {field} present in blinded header")
    # content: original category tokens as whole-field conclusions should not appear
    # as column names; genuine_decay may appear in free text snippets rarely — flag header only
    for pat in LEAK_PATTERNS[:2]:
        if pat.search(header):
            raise SystemExit(f"leakage pattern in header: {pat.pattern}")
    # ensure private key not copied beside public files incorrectly named
    public_private = OUT_DIR / "rq2_original_labels_private.csv"
    if public_private.exists():
        raise SystemExit(
            "private answer key must live under private/ only; found public path copy"
        )


def write_gitignore(path: Path) -> None:
    path.write_text(
        "# Hold out original labels until independent annotation completes.\n"
        "private/\n"
        "rq2_original_labels_private.csv\n"
        "rq2_second_auditor_labels.csv\n"
        "rq2_disagreement_adjudication.csv\n",
        encoding="utf-8",
    )


def main() -> int:
    rows = load_rows()
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    PRIVATE_DIR.mkdir(parents=True, exist_ok=True)

    blinded = OUT_DIR / "rq2_audit_blinded.csv"
    form = OUT_DIR / "rq2_audit_form.xlsx"
    codebook = OUT_DIR / "rq2_audit_codebook.md"
    instructions = OUT_DIR / "INSTRUCTIONS.md"
    private_key = PRIVATE_DIR / "rq2_original_labels_private.csv"
    # Convenience symlink-like copy name requested in protocol (private dir only)
    private_key_alias = PRIVATE_DIR / "rq2_original_labels_private.csv"

    write_blinded_csv(rows, blinded)
    write_xlsx(rows, form)
    write_codebook(codebook)
    write_instructions(instructions)
    write_private_key(rows, private_key)
    assert private_key == private_key_alias
    write_gitignore(OUT_DIR / ".gitignore")
    assert_no_leakage(blinded)

    # QC: event IDs unique; no label columns
    with blinded.open(encoding="utf-8") as handle:
        brows = list(csv.DictReader(handle))
    assert len(brows) == 121
    assert len({r["event_id"] for r in brows}) == 121
    with private_key.open(encoding="utf-8") as handle:
        prows = list(csv.DictReader(handle))
    assert len(prows) == 121
    assert sum(1 for r in prows if r["is_genuine_decay"] == "True") == 0

    print(f"wrote {blinded}")
    print(f"wrote {form}")
    print(f"wrote {codebook}")
    print(f"wrote {instructions}")
    print(f"wrote {private_key} (PRIVATE)")
    print("leakage_check=PASS")
    print("second_auditor_labels=NOT_AVAILABLE")
    return 0


if __name__ == "__main__":
    sys.exit(main())
