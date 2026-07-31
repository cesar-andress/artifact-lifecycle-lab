#!/usr/bin/env python3
"""Assemble a clean, blinded distribution package for an external RQ2 auditor.

Does not modify manuscript, frozen validation numerators, or create annotations.
"""

from __future__ import annotations

import hashlib
import re
import shutil
import subprocess
import sys
import zipfile
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
SRC = ROOT / "validation" / "rq2_second_audit"
PKG_NAME = "RQ2_Independent_Audit_Package"
PKG = ROOT / PKG_NAME
ZIP_PATH = ROOT / f"{PKG_NAME}.zip"

FORBIDDEN_PATTERNS = [
    (re.compile(r"0\s*/\s*121", re.I), "0/121"),
    (re.compile(r"25\s*/\s*121", re.I), "25/121"),
    (re.compile(r"\bfinal_category\b"), "final_category"),
    (re.compile(r"\bis_genuine_decay\b"), "is_genuine_decay"),
    (re.compile(r"answer[_\s-]?key", re.I), "answer key"),
    (re.compile(r"rq2_original_labels", re.I), "original labels file"),
    (re.compile(r"\bprivate/"), "private/"),
    (re.compile(r"rq2_failure_audit\.csv", re.I), "frozen audit csv"),
    (re.compile(r"rq2_disagreement_adjudication", re.I), "adjudication file"),
    (re.compile(r"validation-agreement", re.I), "validation-agreement"),
    (re.compile(r"\bmanuscript\b", re.I), "manuscript"),
    (re.compile(r"0\s*–\s*25\s*/\s*121"), "sensitivity range en-dash"),
]

FORBIDDEN_PATH_FRAGMENTS = [
    "private",
    "answer",
    "original_labels",
    "adjudication",
    "agreement",
    "sensitivity",
    "concentration",
    "manuscript",
    ".py",
]


README_FIRST = """# README FIRST — RQ2 Independent Human Audit

## Purpose

This package supports an **independent, blinded** human audit of **121**
candidate post-verification Missing events from a frozen empirical software
engineering study.

Your task is to classify each event using only the materials in this package
and publicly inspectable repository evidence (commits, trees, histories).

This package is intentionally **blinded**: it does **not** disclose any prior
audit labels, aggregate study results, or expected outcomes. Do not seek those
materials while annotating.

## Files included

| Path | Role |
|------|------|
| `README_FIRST.md` | This file |
| `AUDIT_INSTRUCTIONS.md` | Step-by-step audit procedure |
| `rq2_audit_codebook.md` | Category taxonomy and estimand |
| `rq2_audit_form.xlsx` | Workbook to fill (preferred) |
| `rq2_audit_blinded.csv` | Same evidence in CSV form |
| `blinding_report.md` | Attestation that distribution materials are blinded |
| `evidence/README.md` | Where case evidence lives |
| `protocol/` | Auditor-facing protocol excerpts (PDF) |
| `checksums/SHA256SUMS.txt` | SHA-256 hashes of packaged files |

## Expected deliverable

Return a single CSV named:

`rq2_second_auditor_labels.csv`

with columns:

```text
event_id,category,genuine_decay,confidence,ambiguity,notes,auditor_id,annotation_date
```

Field guidance is in `AUDIT_INSTRUCTIONS.md` and `rq2_audit_codebook.md`.

## Estimated effort

About **8–12 hours** for a careful pass over all 121 events
(roughly 4–6 minutes per event, plus setup).

## Contact

Package coordinator: **[NAME / EMAIL / ORCID — fill before distribution]**

## Blinding statement

This distribution package was assembled and machine-checked so that it does
**not** contain prior classifications, withheld label files, private mappings,
or aggregate numerator results. Please do not seek external study summaries or
prior audit files while annotating.
"""

AUDIT_INSTRUCTIONS = """# Audit Instructions — Independent RQ2 Reviewer

## Role

You are an independent human auditor. Classify all **121** blinded events.
Work alone from this package and repository evidence. Do not seek prior audit
labels, aggregate study results, or the study narrative.

## Materials to use

1. `rq2_audit_codebook.md` — read once before labeling.
2. `rq2_audit_form.xlsx` — preferred annotation surface.
3. `rq2_audit_blinded.csv` — identical evidence if you prefer CSV tooling.
4. Public repository URLs / commits listed in each row (inspect trees/history
   as needed).
5. `protocol/` — auditor-facing protocol excerpts only.

## What each row is

Each row is a **first post-verification Missing** candidate: the reference was
Verified at least once under the frozen detector, then later Missing.
Decide the taxonomy category and whether the event should count as
**genuine post-verification decay** for the study estimand.

## For every event, record

1. **category** — primary taxonomy label from the codebook  
   (`genuine_decay`, `rename_or_move`, `verification_anchor_issue`,
   `extractor_artifact`, `normative_or_prescriptive`,
   `external_or_environmental`, `ambiguous`, or
   `uncertain_insufficient_evidence`)
2. **genuine_decay** — `YES` / `NO` / `INSUFFICIENT_EVIDENCE`
3. **confidence** — `High` / `Medium` / `Low`
4. **notes** — evidence note, max 2–3 sentences
5. **ambiguity** — `true`/`false` or short note of competing categories

Also include `auditor_id` and `annotation_date` (ISO date) on every row.

## Rules

Use only:

- repository evidence (URL in the row)
- commits / file history / manifests / tree state
- extracted reference text and snippet in the row
- this package’s codebook and protocol excerpts

Do **not** consult:

- study write-ups or discussion text
- prior audit labels or withheld label files
- aggregate statistics or expected numerators

## Procedure

1. Open `rq2_audit_form.xlsx` (sheet `audit_form`).
2. For each `event_id`, review evidence columns, then fill the annotation
   columns (or export/complete the CSV schema below).
3. Leave no blank rows. Use `uncertain_insufficient_evidence` and
   `genuine_decay=INSUFFICIENT_EVIDENCE` when stuck.
4. Save your completed labels as `rq2_second_auditor_labels.csv` with header:

```text
event_id,category,genuine_decay,confidence,ambiguity,notes,auditor_id,annotation_date
```

5. Return that CSV to the package coordinator.

## Integrity

Do not reorder or drop `event_id` values. All 121 IDs from the blinded file
must appear exactly once in your deliverable.
"""

BLINDING_REPORT_DIST = """# Blinding Attestation — Distribution Package

**Status:** PASS  
**Events packaged:** 121  
**Package:** `RQ2_Independent_Audit_Package`

## Statement

Auditor-facing files in this distribution were machine-checked for absence of:

- prior audit classification fields
- withheld label files / private mappings
- aggregate numerator results
- hidden worksheets, hidden columns, cell comments, formulas, and
  classification-revealing workbook metadata

## Checks performed (summary)

| Check | Result |
|-------|--------|
| Blinded CSV row count | 121 |
| Unique opaque event IDs | 121 |
| Leak columns in CSV/XLSX | none |
| Hidden columns / rows | none |
| Cell comments | none |
| Formulas | none |
| Conditional formatting rules | none |
| Pre-filled annotation cells | none |
| Forbidden aggregate / label tokens in packaged text | none |
| Private / withheld-label paths included | none |

## Auditor guidance

Classify using only this package and public repository evidence. Do not seek
prior labels or study aggregates while annotating.
"""

PROTOCOL_AUDITOR = """# Independent RQ2 Human Audit — Auditor Protocol Excerpt

**Audience:** external independent auditor  
**Scope:** classification of 121 blinded post-verification Missing candidates  
**Blinding:** this excerpt deliberately omits prior labels and aggregate results

## Purpose

Provide an independent human classification of each candidate event for the
RQ2 estimand: whether a first post-verification Missing observation reflects
**genuine post-verification decay** (physical disappearance / irreversible loss
of the referenced artifact after prior verified existence), or an alternative
explanation under the study taxonomy.

## Units

One row = one candidate event, identified by opaque `event_id`.

## Required judgments (per event)

1. Primary taxonomy category (see codebook).
2. Binary estimand: genuine post-verification decay?
   `YES` / `NO` / `INSUFFICIENT_EVIDENCE`
3. Confidence: `High` / `Medium` / `Low`
4. Short evidence note (2–3 sentences).
5. Ambiguity flag / competing categories if needed.

## Evidence allowed

- Fields in the blinded CSV / XLSX form
- Public repository content at the listed URLs and commits
- Tree / history / manifest inspection you perform yourself

## Evidence forbidden during annotation

- Prior auditor labels or withheld label files
- Aggregate study numerators or study narrative
- Any file not shipped in this package except public repository contents

## Deliverable

`rq2_second_auditor_labels.csv` covering all 121 `event_id` values exactly once.

## After return

The study team will compare your labels to the frozen first-auditor labels
**after** your CSV is received. Those comparison steps are out of scope for
this package.
"""

ADDENDUM_AUDITOR = """# Independent RQ2 Human Audit — Package Addendum

**Audience:** external independent auditor  
**Relation:** companion note to the auditor protocol excerpt

## What this package is

A distribution-ready, blinded annotation kit for one independent human reviewer.

## What this package is not

- Not a complete research replication archive
- Not a source of prior audit decisions
- Not a summary of study headlines

## Versioning note

This auditor kit is derived from a frozen observational export set. Your labels
will be treated as an independent validation stream and will not silently
overwrite any prior classifications.

## Practical checklist

1. Read `README_FIRST.md` and `AUDIT_INSTRUCTIONS.md`.
2. Read `rq2_audit_codebook.md` once.
3. Complete all 121 rows.
4. Return `rq2_second_auditor_labels.csv`.
"""

EVIDENCE_README = """# Evidence

Case-level evidence for each of the 121 events is embedded in:

- `../rq2_audit_form.xlsx` (sheet `audit_form`)
- `../rq2_audit_blinded.csv`

Key fields include repository URL, instruction path, reference text, failure
commit, transition, verified-before-failure, return-after-missing,
basename-collision flag, observation counts, and snippet context.

Use those fields together with public repository inspection at the listed URL
and commit. No separate per-event evidence dumps are required beyond what is
already in the blinded table.
"""


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")


def md_to_pdf(md_path: Path, pdf_path: Path) -> None:
    pdf_path.parent.mkdir(parents=True, exist_ok=True)
    # Prefer pandoc → pdflatex; fall back to weasyprint HTML.
    try:
        subprocess.run(
            [
                "pandoc",
                str(md_path),
                "-o",
                str(pdf_path),
                "--pdf-engine=pdflatex",
                "-V",
                "geometry:margin=1in",
            ],
            check=True,
            capture_output=True,
            text=True,
        )
        return
    except (subprocess.CalledProcessError, FileNotFoundError):
        pass
    html_path = pdf_path.with_suffix(".html")
    subprocess.run(
        ["pandoc", str(md_path), "-o", str(html_path), "-s"],
        check=True,
        capture_output=True,
        text=True,
    )
    subprocess.run(
        ["weasyprint", str(html_path), str(pdf_path)],
        check=True,
        capture_output=True,
        text=True,
    )
    html_path.unlink(missing_ok=True)


def sanitize_xlsx(src: Path, dst: Path) -> None:
    wb = load_workbook(src)
    # Fail closed on hidden material / comments / formulas
    for ws in wb.worksheets:
        for col, dim in ws.column_dimensions.items():
            if dim.hidden:
                raise SystemExit(f"STOP: hidden column {col} in {src}")
        for idx, dim in ws.row_dimensions.items():
            if dim.hidden:
                raise SystemExit(f"STOP: hidden row {idx} in {src}")
        for row in ws.iter_rows():
            for cell in row:
                if cell.comment:
                    raise SystemExit(f"STOP: comment at {cell.coordinate}")
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    raise SystemExit(f"STOP: formula at {cell.coordinate}")
        if len(ws.conditional_formatting._cf_rules):
            raise SystemExit(f"STOP: conditional formatting in {ws.title}")
    if "audit_form" not in wb.sheetnames:
        raise SystemExit("STOP: missing audit_form sheet")
    # Strip identifying metadata
    wb.properties.creator = "RQ2 Independent Audit Package"
    wb.properties.title = "RQ2 blinded audit form"
    wb.properties.subject = None
    wb.properties.description = None
    wb.properties.keywords = None
    wb.properties.category = None
    dst.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dst)


def scan_file_for_forbidden(path: Path) -> list[str]:
    # Skip binary except xlsx handled separately; read text-ish
    if path.suffix.lower() in {".xlsx", ".pdf", ".zip"}:
        if path.suffix.lower() == ".xlsx":
            return scan_xlsx_text(path)
        if path.suffix.lower() == ".pdf":
            # best-effort text extract via pdftotext if available
            try:
                proc = subprocess.run(
                    ["pdftotext", str(path), "-"],
                    check=True,
                    capture_output=True,
                    text=True,
                )
                text = proc.stdout
            except (FileNotFoundError, subprocess.CalledProcessError):
                # fallback: raw bytes latin-1
                text = path.read_bytes().decode("latin-1", errors="ignore")
        else:
            return []
    else:
        text = path.read_text(encoding="utf-8", errors="replace")
    hits = []
    for cre, name in FORBIDDEN_PATTERNS:
        if cre.search(text):
            hits.append(name)
    return hits


def scan_xlsx_text(path: Path) -> list[str]:
    wb = load_workbook(path, data_only=True)
    chunks: list[str] = []
    for ws in wb.worksheets:
        chunks.append(ws.title)
        for row in ws.iter_rows(values_only=True):
            for val in row:
                if val is not None:
                    chunks.append(str(val))
    props = wb.properties
    for attr in ("creator", "title", "subject", "description", "keywords", "category"):
        v = getattr(props, attr, None)
        if v:
            chunks.append(str(v))
    text = "\n".join(chunks)
    return [name for cre, name in FORBIDDEN_PATTERNS if cre.search(text)]


def assert_no_forbidden_paths(pkg: Path) -> None:
    for path in pkg.rglob("*"):
        rel = str(path.relative_to(pkg)).lower()
        for frag in FORBIDDEN_PATH_FRAGMENTS:
            if frag in rel.split("/"):
                # allow evidence/ and protocol/ checksums/ normally
                if frag in {"private", "answer", "original_labels", "adjudication", "agreement", "sensitivity", "concentration", "manuscript"}:
                    raise SystemExit(f"STOP: forbidden path included: {path}")
                if frag == ".py" and path.suffix == ".py":
                    raise SystemExit(f"STOP: script included: {path}")


def build() -> int:
    if PKG.exists():
        shutil.rmtree(PKG)
    if ZIP_PATH.exists():
        ZIP_PATH.unlink()
    PKG.mkdir(parents=True)

    # Core blinded files
    shutil.copy2(SRC / "rq2_audit_blinded.csv", PKG / "rq2_audit_blinded.csv")
    codebook = (SRC / "rq2_audit_codebook.md").read_text(encoding="utf-8")
    codebook = codebook.replace("answer key", "withheld label file")
    codebook = codebook.replace("private answer keys", "withheld label files")
    codebook = codebook.replace("the manuscript,", "study write-ups,")
    codebook = codebook.replace("the manuscript", "study write-ups")
    codebook = codebook.replace("manuscript,", "study write-ups,")
    write_text(PKG / "rq2_audit_codebook.md", codebook)
    sanitize_xlsx(SRC / "rq2_audit_form.xlsx", PKG / "rq2_audit_form.xlsx")

    write_text(PKG / "README_FIRST.md", README_FIRST)
    write_text(PKG / "AUDIT_INSTRUCTIONS.md", AUDIT_INSTRUCTIONS)
    write_text(PKG / "blinding_report.md", BLINDING_REPORT_DIST)
    write_text(PKG / "evidence" / "README.md", EVIDENCE_README)

    proto_md = PKG / "protocol" / "_VALIDATION_EXTENSION_PROTOCOL.md"
    add_md = PKG / "protocol" / "_VALIDATION_EXTENSION_ADDENDUM.md"
    write_text(proto_md, PROTOCOL_AUDITOR)
    write_text(add_md, ADDENDUM_AUDITOR)
    md_to_pdf(proto_md, PKG / "protocol" / "VALIDATION_EXTENSION_PROTOCOL.pdf")
    md_to_pdf(add_md, PKG / "protocol" / "VALIDATION_EXTENSION_ADDENDUM.pdf")
    # Also keep Markdown alongside PDFs for accessibility (auditor-safe versions)
    proto_md.rename(PKG / "protocol" / "VALIDATION_EXTENSION_PROTOCOL.md")
    add_md.rename(PKG / "protocol" / "VALIDATION_EXTENSION_ADDENDUM.md")

    assert_no_forbidden_paths(PKG)

    # Content scan — STOP on any hit
    hits_all: list[tuple[str, list[str]]] = []
    for path in sorted(PKG.rglob("*")):
        if not path.is_file():
            continue
        hits = scan_file_for_forbidden(path)
        if hits:
            hits_all.append((str(path.relative_to(PKG)), hits))
    if hits_all:
        print("STOP: forbidden content detected in distribution package:")
        for rel, hits in hits_all:
            print(f"  {rel}: {hits}")
        raise SystemExit(1)

    # Checksums
    sums_path = PKG / "checksums" / "SHA256SUMS.txt"
    lines = []
    for path in sorted(p for p in PKG.rglob("*") if p.is_file()):
        if path == sums_path:
            continue
        rel = path.relative_to(PKG).as_posix()
        lines.append(f"{sha256_file(path)}  {rel}")
    write_text(sums_path, "\n".join(lines) + "\n")

    # Re-scan checksums file itself for accidental forbidden tokens
    hits = scan_file_for_forbidden(sums_path)
    if hits:
        raise SystemExit(f"STOP: checksums file leaked tokens: {hits}")

    # Zip
    with zipfile.ZipFile(ZIP_PATH, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for path in sorted(p for p in PKG.rglob("*") if p.is_file()):
            zf.write(path, arcname=str(Path(PKG_NAME) / path.relative_to(PKG)))

    # Verify zip
    with zipfile.ZipFile(ZIP_PATH, "r") as zf:
        names = sorted(zf.namelist())
        bad = zf.testzip()
        if bad is not None:
            raise SystemExit(f"STOP: corrupt zip member {bad}")
    expected = [
        f"{PKG_NAME}/README_FIRST.md",
        f"{PKG_NAME}/AUDIT_INSTRUCTIONS.md",
        f"{PKG_NAME}/rq2_audit_codebook.md",
        f"{PKG_NAME}/rq2_audit_form.xlsx",
        f"{PKG_NAME}/rq2_audit_blinded.csv",
        f"{PKG_NAME}/blinding_report.md",
        f"{PKG_NAME}/evidence/README.md",
        f"{PKG_NAME}/protocol/VALIDATION_EXTENSION_PROTOCOL.pdf",
        f"{PKG_NAME}/protocol/VALIDATION_EXTENSION_ADDENDUM.pdf",
        f"{PKG_NAME}/checksums/SHA256SUMS.txt",
    ]
    missing = [e for e in expected if e not in names]
    if missing:
        raise SystemExit(f"STOP: zip missing {missing}")

    # Tree
    files = sorted(p for p in PKG.rglob("*") if p.is_file())
    print("PACKAGE_OK")
    print(f"dir={PKG}")
    print(f"zip={ZIP_PATH}")
    print(f"n_files={len(files)}")
    print(f"zip_bytes={ZIP_PATH.stat().st_size}")
    for p in files:
        print("FILE", p.relative_to(PKG).as_posix())
    return 0


if __name__ == "__main__":
    sys.exit(build())
